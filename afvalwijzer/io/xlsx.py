import logging
from collections.abc import Iterable, Iterator, Callable
from itertools import starmap
from operator import itemgetter
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from afvalwijzer.models import Brongegeven

logger = logging.getLogger(__name__)


def read(file_in: str | Path, filters: dict[str, bool | int | str],
         ) -> Iterator[Brongegeven]:
    """Leest de Afvalwijzer data-export vanuit PowerBI (xlsx).
    """
    def filters_fcn() -> Callable[[list[str]], bool]:
        """Filtert rijen uit de csv op basis van `filters`.
        """
        key = itemgetter(*(
            Brongegeven._fields.index(fld)
            for fld in filters.keys()
        ))
        val = tuple(filters.values())

        def func(x: list[str]) -> bool: return key(x) == val

        return func

    def parse_header(val: str) -> str:
        return val.split('[')[-1].strip(']').lower()

    wb = load_workbook(file_in)
    reader = iter(wb.active.values)

    header = next(reader)
    header = tuple(parse_header(v) for v in header)

    if header != Brongegeven._fields:
        try:
            index = [header.index(fld) for fld in Brongegeven._fields]
        except ValueError:
            logger.warning(header)
            logger.warning(Brongegeven._fields)
            raise ValueError('De header van het xslx-bestand wordt niet herkend.')
        else:
            reader = map(itemgetter(*index), reader)

    if filters:
        reader = filter(filters_fcn(), reader)

    for record in starmap(Brongegeven, reader):
        yield record

    wb.close()


def write(file_out: str | Path, data: Iterable[Brongegeven],
          filters: dict[str, bool | int | str]) -> None:
    """Schrijft de ruwe Afvalwijzer brongegevens naar het xlsx-bestand.
    """
    wb = Workbook()
    ws = wb.active

    ws.append(tuple(header.capitalize() for header in Brongegeven._fields))
    for row in data:
        ws.append(row)

    tbl = Table(displayName='Tabel1', ref=ws.dimensions)
    style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True,
                           showColumnStripes=False)
    tbl.tableStyleInfo = style

    ws.add_table(tbl)
    wb.save(file_out)
